from openpyxl import load_workbook
book = "номенклатура.xlsx"
wb = load_workbook(book)
ws = wb['Microinvest']
ws['K3'] = int(input())

while ws['K3'].value != '777':
    ws['K3'] = str(input())
    for i in range(4, 2249):
        if str(ws['K3'].value) == str(ws['D' + str(i)].value):
            ws['K' + str(i)].value += 1
            a = ws['H' + str(i)].value
            print('+1', a)
            break

wb.save(book)
wb.close()